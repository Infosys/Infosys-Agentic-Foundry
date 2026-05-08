# © 2024-25 Infosys Limited, Bangalore, India. All Rights Reserved.
"""
Token-usage Excel export endpoint.

GET /token-usage/export
    Optional query params: user_id, agent_id, agent_name, date_from, date_to, model

Returns a multi-sheet Excel workbook (.xlsx) as a downloadable attachment.

Sheets
------
1. Summary            – aggregate KPIs (total tokens, total cost, call counts …)
2. Query Usage        – one row per user query  (from query_token_usage table)
3. LLM Call Details   – one row per individual LLM call (from token_usage_logs table)
4. Daily Trend        – tokens & cost aggregated by calendar day  + embedded LineChart
5. Cost by Model      – cost & token breakdown per model           + embedded BarChart
6. Cost by Category   – cost breakdown per call_category           + embedded PieChart
"""

import io
import json
from datetime import datetime, date
from typing import Optional, List, Dict, Any

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

from src.database.repositories import QueryTokenUsageRepository, TokenUsageLogsRepository
from src.api.dependencies import ServiceProvider
from src.auth.dependencies import get_current_user
from src.auth.models import User
from telemetry_wrapper import logger as log


router = APIRouter(prefix="/token-usage", tags=["Token Usage Report"])

# ─────────────────────────────────────────────────────────────────────────────
# Colour palette
# ─────────────────────────────────────────────────────────────────────────────
_HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue
_ALT_ROW_FILL  = PatternFill("solid", fgColor="D6E4F0")   # light blue
_SUMMARY_FILL  = PatternFill("solid", fgColor="2E75B6")   # medium blue
_HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
_TITLE_FONT    = Font(bold=True, size=14, color="1F4E79")
_BOLD_FONT     = Font(bold=True)
_THIN_BORDER   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _style_header_row(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill   = _HEADER_FILL
        cell.font   = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _style_data_row(ws, row: int, n_cols: int, alt: bool = False) -> None:
    fill = _ALT_ROW_FILL if alt else PatternFill("solid", fgColor="FFFFFF")
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill   = fill
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=False)


def _auto_width(ws, min_w: int = 10, max_w: int = 50) -> None:
    for col in ws.columns:
        width = max(
            (len(str(cell.value)) if cell.value else 0 for cell in col),
            default=min_w,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 2, min_w), max_w)


def _write_headers(ws, headers: List[str], row: int = 1) -> None:
    for ci, h in enumerate(headers, 1):
        ws.cell(row=row, column=ci, value=h)
    _style_header_row(ws, row, len(headers))


# ─────────────────────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────────────────────

def _build_summary_sheet(ws, query_rows: List[Dict], log_rows: List[Dict],
                          filters: Dict[str, Any]) -> None:
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "Token Usage & Cost — Summary Report"
    title_cell.font  = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Applied filters
    ws["A3"] = "Filters Applied"
    ws["A3"].font = _BOLD_FONT
    row = 4
    for k, v in filters.items():
        if v:
            ws.cell(row=row, column=1, value=k)
            ws.cell(row=row, column=2, value=str(v))
            row += 1
    if row == 4:
        ws.cell(row=row, column=1, value="(none)")
        row += 1

    row += 1  # blank separator

    # KPI block
    kpi_header_row = row
    ws.cell(row=kpi_header_row, column=1, value="Metric")
    ws.cell(row=kpi_header_row, column=2, value="Value")
    _style_header_row(ws, kpi_header_row, 2)
    row += 1

    total_prompt     = sum(r.get("prompt_tokens",     0)   for r in query_rows)
    total_completion = sum(r.get("completion_tokens", 0)   for r in query_rows)
    total_cached     = sum(r.get("cached_tokens",     0)   for r in query_rows)
    total_tokens     = sum(r.get("total_tokens",      0)   for r in query_rows)
    total_cost       = sum(float(r.get("total_cost",  0.0)) for r in query_rows)
    total_queries    = len(query_rows)
    query_llm_calls  = sum(r.get("total_llm_calls",  0)   for r in query_rows)
    total_llm_calls  = len(log_rows)
    unique_agents    = len({r.get("agent_name") or r.get("agent_id") for r in query_rows} - {None})
    unique_users     = len({r.get("user_id") for r in query_rows} - {None})

    kpis = [
        ("Total Queries",               total_queries),
        ("Query-Related LLM Calls",     query_llm_calls),
        ("Total LLM Calls",             total_llm_calls),
        ("Total Prompt Tokens",         total_prompt),
        ("Total Completion Tokens",     total_completion),
        ("Total Cached Tokens",         total_cached),
        ("Total Tokens",                total_tokens),
        ("Total Cost (USD)",            f"${total_cost:.6f}"),
        ("Unique Agents",               unique_agents),
        ("Unique Users",                unique_users),
    ]
    for i, (metric, value) in enumerate(kpis):
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=value)
        _style_data_row(ws, row, 2, alt=(i % 2 == 1))
        row += 1

    _auto_width(ws)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20


def _build_query_usage_sheet(ws, rows: List[Dict]) -> None:
    ws.title = "Query Usage"

    headers = [
        "ID", "Created At", "User ID", "Agent ID", "Agent Name",
        "Session ID", "Query",
        "Prompt Tokens", "Completion Tokens", "Cached Tokens", "Total Tokens",
        "Prompt Cost ($)", "Completion Cost ($)", "Cached Cost ($)", "Total Cost ($)",
        "Total LLM Calls",
    ]
    _write_headers(ws, headers, row=1)
    ws.row_dimensions[1].height = 30

    for ri, r in enumerate(rows, 2):
        values = [
            r.get("id"),
            r.get("created_at").strftime("%Y-%m-%d %H:%M:%S") if r.get("created_at") else None,
            r.get("user_id"),
            r.get("agent_id"),
            r.get("agent_name"),
            r.get("session_id"),
            r.get("query"),
            r.get("prompt_tokens", 0),
            r.get("completion_tokens", 0),
            r.get("cached_tokens", 0),
            r.get("total_tokens", 0),
            float(r.get("prompt_cost", 0)),
            float(r.get("completion_cost", 0)),
            float(r.get("cached_cost", 0)),
            float(r.get("total_cost", 0)),
            r.get("total_llm_calls", 0),
        ]
        for ci, v in enumerate(values, 1):
            ws.cell(row=ri, column=ci, value=v)
        _style_data_row(ws, ri, len(headers), alt=(ri % 2 == 0))

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _build_llm_calls_sheet(ws, log_rows: List[Dict], query_rows: List[Dict]) -> None:
    """
    Populate from token_usage_logs if available; otherwise explode the llm_calls
    JSONB column from query_token_usage rows.
    """
    ws.title = "LLM Call Details"

    headers = [
        "Timestamp", "User ID", "Agent ID", "Agent Name", "Session ID",
        "Model", "Prompt Tokens", "Completion Tokens", "Cached Tokens", "Total Tokens",
        "Prompt Cost ($)", "Completion Cost ($)", "Cached Cost ($)", "Total Cost ($)",
        "Status", "Call Category", "Call Sub-Category", "Call Operation",
        "Tool Name", "Agent Type", "Agent Component",
    ]
    _write_headers(ws, headers, row=1)
    ws.row_dimensions[1].height = 30

    ri = 2
    if log_rows:
        for r in log_rows:
            values = [
                r.get("timestamp").strftime("%Y-%m-%d %H:%M:%S") if r.get("timestamp") else None,
                r.get("user_id"),
                str(r.get("agent_id", "")),
                r.get("agent_name"),
                r.get("session_id"),
                r.get("model_name"),
                r.get("prompt_tokens", 0),
                r.get("completion_tokens", 0),
                r.get("cached_tokens", 0),
                r.get("total_tokens", 0),
                float(r.get("prompt_tokens_cost", 0) or 0),
                float(r.get("completion_tokens_cost", 0) or 0),
                float(r.get("cached_tokens_cost", 0) or 0),
                float(r.get("total_cost", 0) or 0),
                r.get("status"),
                r.get("call_category"),
                r.get("call_sub_category"),
                r.get("call_operation"),
                r.get("tool_name"),
                r.get("agent_type"),
                r.get("agent_component"),
            ]
            for ci, v in enumerate(values, 1):
                ws.cell(row=ri, column=ci, value=v)
            _style_data_row(ws, ri, len(headers), alt=(ri % 2 == 0))
            ri += 1
    else:
        # Fallback: explode llm_calls JSONB from query rows
        for qr in query_rows:
            raw_calls = qr.get("llm_calls") or []
            if isinstance(raw_calls, str):
                try:
                    raw_calls = json.loads(raw_calls)
                except Exception:
                    raw_calls = []
            created_at = qr.get("created_at")
            ts_str = created_at.strftime("%Y-%m-%d %H:%M:%S") if created_at else None
            for call in raw_calls:
                values = [
                    ts_str,
                    qr.get("user_id"),
                    qr.get("agent_id"),
                    qr.get("agent_name"),
                    qr.get("session_id"),
                    call.get("model"),
                    call.get("prompt_tokens", 0),
                    call.get("completion_tokens", 0),
                    call.get("cached_tokens", 0),
                    call.get("total_tokens", 0),
                    float(call.get("prompt_cost", 0) or 0),
                    float(call.get("completion_cost", 0) or 0),
                    float(call.get("cached_cost", 0) or 0),
                    float(call.get("total_cost", 0) or 0),
                    call.get("status"),
                    call.get("call_category"),
                    call.get("call_sub_category"),
                    None, None, None, None,
                ]
                for ci, v in enumerate(values, 1):
                    ws.cell(row=ri, column=ci, value=v)
                _style_data_row(ws, ri, len(headers), alt=(ri % 2 == 0))
                ri += 1

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _build_daily_trend_sheet(ws, log_rows: List[Dict], query_rows: List[Dict]) -> None:
    ws.title = "Daily Trend"

    # Aggregate by day
    day_map: Dict[str, Dict[str, float]] = {}

    def _add_day(day_str: str, tokens: int, cost: float) -> None:
        if day_str not in day_map:
            day_map[day_str] = {"total_tokens": 0, "total_cost": 0.0, "call_count": 0}
        day_map[day_str]["total_tokens"] += tokens
        day_map[day_str]["total_cost"]   += cost
        day_map[day_str]["call_count"]   += 1

    if log_rows:
        for r in log_rows:
            ts = r.get("timestamp")
            if ts:
                day_str = ts.strftime("%Y-%m-%d")
                _add_day(day_str, r.get("total_tokens", 0), float(r.get("total_cost", 0) or 0))
    else:
        for r in query_rows:
            ts = r.get("created_at")
            if ts:
                day_str = ts.strftime("%Y-%m-%d")
                _add_day(day_str, r.get("total_tokens", 0), float(r.get("total_cost", 0) or 0))

    sorted_days = sorted(day_map.keys())

    headers = ["Date", "Total Tokens", "Total Cost ($)", "LLM Call Count"]
    _write_headers(ws, headers, row=1)

    for ri, day in enumerate(sorted_days, 2):
        d = day_map[day]
        ws.cell(row=ri, column=1, value=day)
        ws.cell(row=ri, column=2, value=d["total_tokens"])
        ws.cell(row=ri, column=3, value=round(d["total_cost"], 8))
        ws.cell(row=ri, column=4, value=d["call_count"])
        _style_data_row(ws, ri, 4, alt=(ri % 2 == 0))

    _auto_width(ws)
    ws.freeze_panes = "A2"

    n = len(sorted_days)
    if n < 2:
        return

    # LineChart — tokens over time
    chart = LineChart()
    chart.title  = "Daily Token Usage Trend"
    chart.style  = 10
    chart.y_axis.title = "Tokens"
    chart.x_axis.title = "Date"
    chart.height = 14
    chart.width  = 26

    data_ref   = Reference(ws, min_col=2, min_row=1, max_row=1 + n)
    cats_ref   = Reference(ws, min_col=1, min_row=2, max_row=1 + n)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.line.solidFill = "2E75B6"

    ws.add_chart(chart, f"F2")


def _build_cost_by_model_sheet(ws, log_rows: List[Dict], query_rows: List[Dict]) -> None:
    ws.title = "Cost by Model"

    model_map: Dict[str, Dict[str, float]] = {}

    def _add_model(model: str, tokens: int, cost: float) -> None:
        model = model or "unknown"
        if model not in model_map:
            model_map[model] = {"total_tokens": 0, "total_cost": 0.0, "call_count": 0}
        model_map[model]["total_tokens"] += tokens
        model_map[model]["total_cost"]   += cost
        model_map[model]["call_count"]   += 1

    if log_rows:
        for r in log_rows:
            _add_model(r.get("model_name", ""), r.get("total_tokens", 0),
                       float(r.get("total_cost", 0) or 0))
    else:
        for qr in query_rows:
            raw_calls = qr.get("llm_calls") or []
            if isinstance(raw_calls, str):
                try:
                    raw_calls = json.loads(raw_calls)
                except Exception:
                    raw_calls = []
            for call in raw_calls:
                _add_model(call.get("model", ""), call.get("total_tokens", 0),
                           float(call.get("total_cost", 0) or 0))

    sorted_models = sorted(model_map.keys())

    headers = ["Model", "Total Tokens", "Total Cost ($)", "LLM Call Count"]
    _write_headers(ws, headers, row=1)

    for ri, model in enumerate(sorted_models, 2):
        d = model_map[model]
        ws.cell(row=ri, column=1, value=model)
        ws.cell(row=ri, column=2, value=d["total_tokens"])
        ws.cell(row=ri, column=3, value=round(d["total_cost"], 8))
        ws.cell(row=ri, column=4, value=d["call_count"])
        _style_data_row(ws, ri, 4, alt=(ri % 2 == 0))

    _auto_width(ws)
    ws.freeze_panes = "A2"

    n = len(sorted_models)
    if n < 1:
        return

    # BarChart — cost per model
    chart = BarChart()
    chart.type   = "col"
    chart.title  = "Cost by Model"
    chart.style  = 10
    chart.y_axis.title = "Total Cost (USD)"
    chart.x_axis.title = "Model"
    chart.height = 14
    chart.width  = 26

    data_ref = Reference(ws, min_col=3, min_row=1, max_row=1 + n)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=1 + n)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "2E75B6"

    ws.add_chart(chart, "F2")


def _build_cost_by_category_sheet(ws, log_rows: List[Dict], query_rows: List[Dict]) -> None:
    ws.title = "Cost by Category"

    cat_map: Dict[str, Dict[str, float]] = {}

    def _add_cat(cat: str, tokens: int, cost: float) -> None:
        cat = cat or "uncategorized"
        if cat not in cat_map:
            cat_map[cat] = {"total_tokens": 0, "total_cost": 0.0, "call_count": 0}
        cat_map[cat]["total_tokens"] += tokens
        cat_map[cat]["total_cost"]   += cost
        cat_map[cat]["call_count"]   += 1

    if log_rows:
        for r in log_rows:
            _add_cat(r.get("call_category", ""), r.get("total_tokens", 0),
                     float(r.get("total_cost", 0) or 0))
    else:
        for qr in query_rows:
            raw_calls = qr.get("llm_calls") or []
            if isinstance(raw_calls, str):
                try:
                    raw_calls = json.loads(raw_calls)
                except Exception:
                    raw_calls = []
            for call in raw_calls:
                _add_cat(call.get("call_category", ""), call.get("total_tokens", 0),
                         float(call.get("total_cost", 0) or 0))

    sorted_cats = sorted(cat_map.keys())

    headers = ["Category", "Total Tokens", "Total Cost ($)", "LLM Call Count"]
    _write_headers(ws, headers, row=1)

    for ri, cat in enumerate(sorted_cats, 2):
        d = cat_map[cat]
        ws.cell(row=ri, column=1, value=cat)
        ws.cell(row=ri, column=2, value=d["total_tokens"])
        ws.cell(row=ri, column=3, value=round(d["total_cost"], 8))
        ws.cell(row=ri, column=4, value=d["call_count"])
        _style_data_row(ws, ri, 4, alt=(ri % 2 == 0))

    _auto_width(ws)
    ws.freeze_panes = "A2"

    n = len(sorted_cats)
    if n < 1:
        return

    # PieChart — cost share by category
    chart = PieChart()
    chart.title  = "Cost Share by Call Category"
    chart.style  = 10
    chart.height = 14
    chart.width  = 20

    data_ref = Reference(ws, min_col=3, min_row=1, max_row=1 + n)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=1 + n)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.dataLabels = None

    ws.add_chart(chart, "F2")


# ─────────────────────────────────────────────────────────────────────────────
# Endpoint
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/export",
    summary="Download token-usage & cost report as Excel",
    response_description="Excel workbook (.xlsx) with 6 analysis sheets and embedded charts",
)
async def export_token_usage_report(
    user_id: Optional[str]   = Query(None, description="Filter by user e-mail"),
    agent_id: Optional[str]  = Query(None, description="Filter by agent ID"),
    agent_name: Optional[str] = Query(None, description="Filter by agent name (partial match)"),
    date_from: Optional[date] = Query(None, description="Start date  (YYYY-MM-DD)"),
    date_to: Optional[date]   = Query(None, description="End date    (YYYY-MM-DD)"),
    model: Optional[str]      = Query(None, description="Filter by model name (partial match, applied to LLM call details)"),
    user_data: User = Depends(get_current_user),
    query_token_usage_repo: QueryTokenUsageRepository = Depends(ServiceProvider.get_query_token_usage_repo),
    token_usage_logs_repo: TokenUsageLogsRepository   = Depends(ServiceProvider.get_token_usage_logs_repo),
):
    """
    Export a token-usage and cost report as an Excel (.xlsx) file.

    The workbook contains 6 sheets:

    | Sheet              | Content                                    |
    |--------------------|--------------------------------------------|
    | Summary            | Aggregate KPIs for the selected filters    |
    | Query Usage        | One row per user query                     |
    | LLM Call Details   | One row per individual LLM call            |
    | Daily Trend        | Tokens & cost by day + line chart          |
    | Cost by Model      | Cost breakdown per model + bar chart       |
    | Cost by Category   | Cost breakdown per category + pie chart    |
    """
    dt_from = datetime(date_from.year, date_from.month, date_from.day) if date_from else None
    dt_to   = datetime(date_to.year,   date_to.month,   date_to.day,
                       23, 59, 59) if date_to else None

    log.info(
        f"[TokenUsageExport] Request by {user_data.email} | "
        f"filters: user_id={user_id}, agent_id={agent_id}, agent_name={agent_name}, "
        f"date_from={date_from}, date_to={date_to}, model={model}"
    )

    # Fetch data from both sources concurrently
    import asyncio
    query_rows, log_rows = await asyncio.gather(
        query_token_usage_repo.get_report_data(
            user_id=user_id,
            agent_id=agent_id,
            agent_name=agent_name,
            date_from=dt_from,
            date_to=dt_to,
        ),
        token_usage_logs_repo.get_report_data(
            user_id=user_id,
            agent_id=agent_id,
            agent_name=agent_name,
            date_from=dt_from,
            date_to=dt_to,
            model=model,
        ),
    )

    log.info(
        f"[TokenUsageExport] Fetched {len(query_rows)} query rows, "
        f"{len(log_rows)} LLM call rows"
    )

    if not query_rows and not log_rows:
        raise HTTPException(
            status_code=404,
            detail="No token usage data found for the given filters.",
        )

    filters = {
        "User ID":    user_id,
        "Agent ID":   agent_id,
        "Agent Name": agent_name,
        "Date From":  str(date_from) if date_from else None,
        "Date To":    str(date_to)   if date_to   else None,
        "Model":      model,
    }

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    _build_summary_sheet(wb.create_sheet(),       query_rows, log_rows, filters)
    _build_query_usage_sheet(wb.create_sheet(),   query_rows)
    _build_llm_calls_sheet(wb.create_sheet(),     log_rows, query_rows)
    _build_daily_trend_sheet(wb.create_sheet(),   log_rows, query_rows)
    _build_cost_by_model_sheet(wb.create_sheet(), log_rows, query_rows)
    _build_cost_by_category_sheet(wb.create_sheet(), log_rows, query_rows)

    # Serialise to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"token_usage_report_{ts}.xlsx"

    log.info(f"[TokenUsageExport] Workbook built — returning '{filename}'")

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
